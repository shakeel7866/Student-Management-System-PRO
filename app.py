# -*- coding: utf-8 -*-
"""
Student Management System PRO — Flask Web Edition
====================================================
Responsive (mobile + desktop) web app with:
  - Role-based login: Admin, Teacher, Student, Parent
  - Admin panel: create accounts, choose display density (Small / Medium)
  - Students: scrollable list, add/edit/delete, search
  - Subjects: add / rename / delete
  - Attendance tracking
  - Exams & Results: scrollable table with up/down scroll buttons
  - Fees management
  - Teacher Management: profiles, subject assignment, salary records
  - Parent Portal: view children's attendance/results, message teachers
  - Export to Excel and PDF
  - Urdu / English language toggle
  - Works on both mobile and desktop browsers (Bootstrap 5 responsive layout)

Run:
    pip install flask openpyxl reportlab
    python app.py
Then open http://127.0.0.1:5000
Default admin login created on first run: admin / admin123
"""

import os
import io
import sqlite3
import datetime
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for, session,
    flash, g, send_file, abort
)
from werkzeug.security import generate_password_hash, check_password_hash

from translations import get_translator

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "school.db")
DEFAULT_SUBJECTS = ["Math", "English", "Science", "Urdu", "Computer Science", "Physics", "Chemistry"]

app = Flask(__name__)
app.secret_key = "change-this-secret-key-in-production"

ROLES = ["admin", "teacher", "student", "parent"]


# ─────────────────────────────────────────────────────────────────────────────
#  DATABASE
# ─────────────────────────────────────────────────────────────────────────────
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exc=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    first_time = not os.path.exists(DB_PATH)
    db = sqlite3.connect(DB_PATH)
    db.execute("""CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        full_name TEXT,
        role TEXT NOT NULL,
        density TEXT DEFAULT 'medium',
        created_at TEXT
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS subjects(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        position INTEGER DEFAULT 0
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS students(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        roll_no TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        father TEXT,
        email TEXT,
        class_name TEXT,
        age INTEGER,
        attendance REAL DEFAULT 100,
        student_user_id INTEGER,
        parent_user_id INTEGER,
        admission_fee REAL DEFAULT 5000,
        monthly_fee REAL DEFAULT 2000,
        fee_paid REAL DEFAULT 0,
        fee_pending REAL DEFAULT 7000,
        fee_status TEXT DEFAULT 'unpaid'
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS marks(
        student_id INTEGER NOT NULL,
        subject_id INTEGER NOT NULL,
        marks REAL DEFAULT 0,
        PRIMARY KEY (student_id, subject_id),
        FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE CASCADE,
        FOREIGN KEY(subject_id) REFERENCES subjects(id) ON DELETE CASCADE
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS teachers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER UNIQUE,
        name TEXT NOT NULL,
        phone TEXT,
        salary REAL DEFAULT 0,
        subject_ids TEXT DEFAULT ''
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS messages(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sender_id INTEGER,
        recipient_id INTEGER,
        student_id INTEGER,
        body TEXT,
        created_at TEXT
    )""")
    db.commit()

    if first_time:
        # Seed default admin + subjects
        ph = generate_password_hash("admin123")
        db.execute(
            "INSERT INTO users(username,password_hash,full_name,role,density,created_at) VALUES (?,?,?,?,?,?)",
            ("admin", ph, "System Admin", "admin", "medium", datetime.datetime.now().isoformat()),
        )
        for i, s in enumerate(DEFAULT_SUBJECTS):
            db.execute("INSERT INTO subjects(name,position) VALUES (?,?)", (s, i))
        db.commit()
    db.close()


# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    db = get_db()
    return db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()


def login_required(roles=None):
    def deco(fn):
        @wraps(fn)
        def wrapper(*a, **kw):
            user = current_user()
            if not user:
                return redirect(url_for("login"))
            if roles and user["role"] not in roles:
                abort(403)
            return fn(*a, **kw)
        return wrapper
    return deco


def grade_for(m):
    if m >= 90: return "A+"
    if m >= 80: return "A"
    if m >= 70: return "B"
    if m >= 60: return "C"
    if m >= 50: return "D"
    return "F"


def get_subjects():
    db = get_db()
    return db.execute("SELECT * FROM subjects ORDER BY position, id").fetchall()


def get_student_marks(student_id):
    db = get_db()
    rows = db.execute("""
        SELECT s.id as subject_id, s.name, COALESCE(m.marks,0) as marks
        FROM subjects s LEFT JOIN marks m
          ON m.subject_id = s.id AND m.student_id = ?
        ORDER BY s.position, s.id
    """, (student_id,)).fetchall()
    return rows


def avg_for_student(student_id):
    rows = get_student_marks(student_id)
    vals = [r["marks"] for r in rows]
    return round(sum(vals) / len(vals), 1) if vals else 0.0


@app.context_processor
def inject_globals():
    lang = session.get("lang", "en")
    return dict(
        t=get_translator(lang),
        lang=lang,
        current_user=current_user(),
        density=(current_user()["density"] if current_user() else session.get("density", "medium")),
    )


# ─────────────────────────────────────────────────────────────────────────────
#  AUTH
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return redirect(url_for("dashboard") if current_user() else url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user():
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        remember = request.form.get("remember")
        db = get_db()
        row = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        if row and check_password_hash(row["password_hash"], password):
            session["user_id"] = row["id"]
            session["density"] = row["density"]
            if remember:
                session.permanent = True
                app.permanent_session_lifetime = datetime.timedelta(days=30)
            return redirect(url_for("dashboard"))
        error = "Invalid username or password."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/lang/<code>")
def set_lang(code):
    if code in ("en", "ur"):
        session["lang"] = code
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/density/<level>")
@login_required()
def set_density(level):
    if level in ("small", "medium"):
        session["density"] = level
        db = get_db()
        db.execute("UPDATE users SET density=? WHERE id=?", (level, session["user_id"]))
        db.commit()
    return redirect(request.referrer or url_for("dashboard"))


# ─────────────────────────────────────────────────────────────────────────────
#  DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/dashboard")
@login_required()
def dashboard():
    user = current_user()
    db = get_db()
    students = db.execute("SELECT * FROM students").fetchall()
    avgs = [avg_for_student(s["id"]) for s in students]
    n = len(students)
    stats = {
        "total_students": n,
        "class_average": round(sum(avgs) / n, 1) if n else 0,
        "pass_rate": round(sum(1 for a in avgs if a >= 50) / n * 100) if n else 0,
        "top_scorer": "",
    }
    if students:
        top_idx = avgs.index(max(avgs))
        stats["top_scorer"] = students[top_idx]["name"]

    my_children = []
    my_subjects = []
    if user["role"] == "parent":
        my_children = db.execute(
            "SELECT * FROM students WHERE parent_user_id=?", (user["id"],)
        ).fetchall()
    if user["role"] == "teacher":
        trow = db.execute("SELECT * FROM teachers WHERE user_id=?", (user["id"],)).fetchone()
        if trow and trow["subject_ids"]:
            ids = [int(x) for x in trow["subject_ids"].split(",") if x]
            my_subjects = db.execute(
                f"SELECT * FROM subjects WHERE id IN ({','.join('?'*len(ids))})", ids
            ).fetchall() if ids else []

    my_profile = None
    if user["role"] == "student":
        my_profile = db.execute("SELECT * FROM students WHERE student_user_id=?", (user["id"],)).fetchone()

    return render_template(
        "dashboard.html", stats=stats, students=students[:8],
        my_children=my_children, my_subjects=my_subjects, my_profile=my_profile,
        avg_for_student=avg_for_student, grade_for=grade_for,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  ADMIN PANEL — create accounts, choose display density
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/admin", methods=["GET", "POST"])
@login_required(roles=["admin"])
def admin_panel():
    db = get_db()
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        full_name = request.form.get("full_name", "").strip()
        role = request.form.get("role", "student")
        density = request.form.get("density", "medium")
        link_roll = request.form.get("link_roll", "").strip()

        if not username or not password or role not in ROLES:
            error = "Username, password and a valid role are required."
        elif db.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
            error = f"Username '{username}' already exists."
        else:
            ph = generate_password_hash(password)
            cur = db.execute(
                "INSERT INTO users(username,password_hash,full_name,role,density,created_at) VALUES (?,?,?,?,?,?)",
                (username, ph, full_name, role, density, datetime.datetime.now().isoformat()),
            )
            new_uid = cur.lastrowid
            if role == "student" and link_roll:
                db.execute("UPDATE students SET student_user_id=? WHERE roll_no=?", (new_uid, link_roll))
            elif role == "parent" and link_roll:
                db.execute("UPDATE students SET parent_user_id=? WHERE roll_no=?", (new_uid, link_roll))
            elif role == "teacher":
                db.execute("INSERT INTO teachers(user_id,name,phone,salary,subject_ids) VALUES (?,?,?,?,?)",
                           (new_uid, full_name or username, "", 0, ""))
            db.commit()
            flash(f"Account '{username}' created as {role}.", "success")
            return redirect(url_for("admin_panel"))

    users = db.execute("SELECT * FROM users ORDER BY id").fetchall()
    students = db.execute("SELECT roll_no, name FROM students ORDER BY name").fetchall()
    return render_template("admin_panel.html", users=users, students=students, error=error)


@app.route("/admin/users/<int:uid>/delete", methods=["POST"])
@login_required(roles=["admin"])
def admin_delete_user(uid):
    if uid == session.get("user_id"):
        flash("You cannot remove your own account while logged in.", "danger")
        return redirect(url_for("admin_panel"))
    db = get_db()
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    db.execute("DELETE FROM teachers WHERE user_id=?", (uid,))
    db.commit()
    flash("Account removed.", "success")
    return redirect(url_for("admin_panel"))


# ─────────────────────────────────────────────────────────────────────────────
#  STUDENTS
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/students")
@login_required(roles=["admin", "teacher"])
def students_list():
    db = get_db()
    q = request.args.get("q", "").strip().lower()
    rows = db.execute("SELECT * FROM students ORDER BY name").fetchall()
    if q:
        rows = [r for r in rows if q in r["name"].lower() or q in r["roll_no"].lower()
                or q in (r["class_name"] or "").lower()]
    enriched = []
    for r in rows:
        am = avg_for_student(r["id"])
        enriched.append({"row": r, "avg": am, "grade": grade_for(am)})
    return render_template("students_list.html", students=enriched, q=q)


@app.route("/students/add", methods=["GET", "POST"])
@login_required(roles=["admin", "teacher"])
def student_add():
    return student_form(None)


@app.route("/students/edit/<int:sid>", methods=["GET", "POST"])
@login_required(roles=["admin", "teacher"])
def student_edit(sid):
    return student_form(sid)


def student_form(sid):
    db = get_db()
    subjects = get_subjects()
    student = None
    marks_map = {}
    if sid:
        student = db.execute("SELECT * FROM students WHERE id=?", (sid,)).fetchone()
        if not student:
            abort(404)
        for m in get_student_marks(sid):
            marks_map[m["subject_id"]] = m["marks"]

    error = None
    if request.method == "POST":
        if request.form.get("action") == "skip":
            flash("Skipped — nothing was saved.", "info")
            return redirect(url_for("students_list"))

        roll_no = request.form.get("roll_no", "").strip()
        name = request.form.get("name", "").strip()
        if not roll_no or not name:
            error = "Roll No. and Name are required."
        else:
            dup = db.execute("SELECT id FROM students WHERE roll_no=? AND id IS NOT ?",
                              (roll_no, sid or -1)).fetchone()
            if dup:
                error = f"Roll No. '{roll_no}' already exists."

        if not error:
            father = request.form.get("father", "")
            email = request.form.get("email", "")
            class_name = request.form.get("class_name", "")
            try:
                age = int(request.form.get("age") or 0)
            except ValueError:
                age = 0
            try:
                attendance = float(request.form.get("attendance") or 100)
            except ValueError:
                attendance = 100
            admission_fee = float(request.form.get("admission_fee") or 0)
            monthly_fee = float(request.form.get("monthly_fee") or 0)
            fee_paid = float(request.form.get("fee_paid") or 0)
            fee_pending = float(request.form.get("fee_pending") or 0)
            fee_status = request.form.get("fee_status", "unpaid")

            if sid:
                db.execute("""UPDATE students SET roll_no=?,name=?,father=?,email=?,class_name=?,age=?,
                              attendance=?,admission_fee=?,monthly_fee=?,fee_paid=?,fee_pending=?,fee_status=?
                              WHERE id=?""",
                           (roll_no, name, father, email, class_name, age, attendance,
                            admission_fee, monthly_fee, fee_paid, fee_pending, fee_status, sid))
                new_id = sid
            else:
                cur = db.execute("""INSERT INTO students
                    (roll_no,name,father,email,class_name,age,attendance,
                     admission_fee,monthly_fee,fee_paid,fee_pending,fee_status)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (roll_no, name, father, email, class_name, age, attendance,
                     admission_fee, monthly_fee, fee_paid, fee_pending, fee_status))
                new_id = cur.lastrowid

            for s in subjects:
                key = f"marks_{s['id']}"
                if key in request.form:
                    try:
                        mval = max(0, min(100, float(request.form.get(key) or 0)))
                    except ValueError:
                        mval = 0
                    db.execute("""INSERT INTO marks(student_id,subject_id,marks) VALUES (?,?,?)
                                  ON CONFLICT(student_id,subject_id) DO UPDATE SET marks=excluded.marks""",
                               (new_id, s["id"], mval))
            db.commit()
            flash(f"Student '{name}' saved.", "success")
            if request.form.get("action") == "save_new":
                return redirect(url_for("student_add"))
            return redirect(url_for("students_list"))

    return render_template("student_form.html", student=student, subjects=subjects,
                            marks_map=marks_map, error=error)


@app.route("/students/delete/<int:sid>", methods=["POST"])
@login_required(roles=["admin"])
def student_delete(sid):
    db = get_db()
    db.execute("DELETE FROM students WHERE id=?", (sid,))
    db.execute("DELETE FROM marks WHERE student_id=?", (sid,))
    db.commit()
    flash("Student deleted.", "success")
    return redirect(url_for("students_list"))


# ─────────────────────────────────────────────────────────────────────────────
#  SUBJECTS
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/subjects", methods=["GET", "POST"])
@login_required(roles=["admin"])
def subjects_page():
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            name = request.form.get("name", "").strip()
            if name:
                exists = db.execute("SELECT 1 FROM subjects WHERE name=?", (name,)).fetchone()
                if exists:
                    flash("Subject already exists.", "warning")
                else:
                    pos = db.execute("SELECT COALESCE(MAX(position),0)+1 FROM subjects").fetchone()[0]
                    db.execute("INSERT INTO subjects(name,position) VALUES (?,?)", (name, pos))
                    db.commit()
                    flash(f"Added subject '{name}'.", "success")
        elif action == "rename":
            sid = request.form.get("subject_id")
            new_name = request.form.get("new_name", "").strip()
            if sid and new_name:
                db.execute("UPDATE subjects SET name=? WHERE id=?", (new_name, sid))
                db.commit()
                flash("Subject renamed.", "success")
        elif action == "delete":
            sid = request.form.get("subject_id")
            db.execute("DELETE FROM subjects WHERE id=?", (sid,))
            db.execute("DELETE FROM marks WHERE subject_id=?", (sid,))
            db.commit()
            flash("Subject deleted.", "success")
        return redirect(url_for("subjects_page"))

    subjects = get_subjects()
    return render_template("subjects.html", subjects=subjects)


# ─────────────────────────────────────────────────────────────────────────────
#  ATTENDANCE
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/attendance", methods=["GET", "POST"])
@login_required(roles=["admin", "teacher"])
def attendance_page():
    db = get_db()
    if request.method == "POST":
        for key, val in request.form.items():
            if key.startswith("att_"):
                sid = key.split("_", 1)[1]
                try:
                    v = max(0, min(100, float(val)))
                except ValueError:
                    v = 100
                db.execute("UPDATE students SET attendance=? WHERE id=?", (v, sid))
        db.commit()
        flash("Attendance saved.", "success")
        return redirect(url_for("attendance_page"))

    students = db.execute("SELECT * FROM students ORDER BY name").fetchall()
    return render_template("attendance.html", students=students)


# ─────────────────────────────────────────────────────────────────────────────
#  EXAMS & RESULTS
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/exams", methods=["GET", "POST"])
@login_required(roles=["admin", "teacher"])
def exams_page():
    db = get_db()
    if request.method == "POST":
        sid = request.form.get("student_id")
        subjects = get_subjects()
        for s in subjects:
            key = f"marks_{s['id']}"
            if key in request.form:
                try:
                    mval = max(0, min(100, float(request.form.get(key) or 0)))
                except ValueError:
                    mval = 0
                db.execute("""INSERT INTO marks(student_id,subject_id,marks) VALUES (?,?,?)
                              ON CONFLICT(student_id,subject_id) DO UPDATE SET marks=excluded.marks""",
                           (sid, s["id"], mval))
        db.commit()
        flash("Marks updated.", "success")
        return redirect(url_for("exams_page"))

    subjects = get_subjects()
    students = db.execute("SELECT * FROM students ORDER BY name").fetchall()
    rows = []
    for st in students:
        marks = {m["subject_id"]: m["marks"] for m in get_student_marks(st["id"])}
        am = avg_for_student(st["id"])
        rows.append({"student": st, "marks": marks, "avg": am, "grade": grade_for(am)})
    rows.sort(key=lambda r: r["avg"], reverse=True)
    return render_template("exams.html", subjects=subjects, rows=rows)


# ─────────────────────────────────────────────────────────────────────────────
#  FEES
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/fees", methods=["GET", "POST"])
@login_required(roles=["admin"])
def fees_page():
    db = get_db()
    if request.method == "POST":
        sid = request.form.get("student_id")
        admission_fee = float(request.form.get("admission_fee") or 0)
        monthly_fee = float(request.form.get("monthly_fee") or 0)
        fee_paid = float(request.form.get("fee_paid") or 0)
        fee_pending = float(request.form.get("fee_pending") or 0)
        fee_status = request.form.get("fee_status", "unpaid")
        db.execute("""UPDATE students SET admission_fee=?,monthly_fee=?,fee_paid=?,
                      fee_pending=?,fee_status=? WHERE id=?""",
                   (admission_fee, monthly_fee, fee_paid, fee_pending, fee_status, sid))
        db.commit()
        flash("Fees updated.", "success")
        return redirect(url_for("fees_page"))

    students = db.execute("SELECT * FROM students ORDER BY name").fetchall()
    totals = {
        "admission": sum(s["admission_fee"] for s in students),
        "monthly": sum(s["monthly_fee"] for s in students),
        "paid": sum(s["fee_paid"] for s in students),
        "pending": sum(s["fee_pending"] for s in students),
    }
    return render_template("fees.html", students=students, totals=totals)


# ─────────────────────────────────────────────────────────────────────────────
#  TEACHER MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/teachers", methods=["GET", "POST"])
@login_required(roles=["admin"])
def teachers_page():
    db = get_db()
    if request.method == "POST":
        tid = request.form.get("teacher_id")
        name = request.form.get("name", "").strip()
        phone = request.form.get("phone", "").strip()
        salary = float(request.form.get("salary") or 0)
        subj_ids = request.form.getlist("subject_ids")
        subj_str = ",".join(subj_ids)
        if tid:
            db.execute("UPDATE teachers SET name=?,phone=?,salary=?,subject_ids=? WHERE id=?",
                       (name, phone, salary, subj_str, tid))
            flash("Teacher updated.", "success")
        db.commit()
        return redirect(url_for("teachers_page"))

    teachers = db.execute("""
        SELECT t.*, u.username FROM teachers t
        LEFT JOIN users u ON u.id = t.user_id
        ORDER BY t.name
    """).fetchall()
    subjects = get_subjects()
    teacher_subjects = {}
    for t in teachers:
        ids = [int(x) for x in t["subject_ids"].split(",") if x]
        teacher_subjects[t["id"]] = ids
    return render_template("teachers.html", teachers=teachers, subjects=subjects,
                            teacher_subjects=teacher_subjects)


# ─────────────────────────────────────────────────────────────────────────────
#  PARENT PORTAL
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/parent", methods=["GET", "POST"])
@login_required(roles=["parent"])
def parent_portal():
    db = get_db()
    user = current_user()
    children = db.execute("SELECT * FROM students WHERE parent_user_id=?", (user["id"],)).fetchall()

    if request.method == "POST":
        student_id = request.form.get("student_id")
        body = request.form.get("body", "").strip()
        if body:
            db.execute("""INSERT INTO messages(sender_id,recipient_id,student_id,body,created_at)
                          VALUES (?,?,?,?,?)""",
                       (user["id"], None, student_id, body, datetime.datetime.now().isoformat()))
            db.commit()
            flash("Message sent to the teacher.", "success")
        return redirect(url_for("parent_portal"))

    child_cards = []
    for c in children:
        am = avg_for_student(c["id"])
        msgs = db.execute("""SELECT m.*, u.full_name as sender_name FROM messages m
                              LEFT JOIN users u ON u.id = m.sender_id
                              WHERE m.student_id=? ORDER BY m.created_at DESC LIMIT 10""",
                           (c["id"],)).fetchall()
        child_cards.append({
            "student": c, "avg": am, "grade": grade_for(am),
            "marks": get_student_marks(c["id"]), "messages": msgs,
        })
    return render_template("parent_portal.html", child_cards=child_cards)


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORTS
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/export/excel")
@login_required(roles=["admin", "teacher"])
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    db = get_db()
    subjects = get_subjects()
    students = db.execute("SELECT * FROM students ORDER BY name").fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    navy = PatternFill("solid", fgColor="1A1F36")
    headers = ["Roll No.", "Name", "Father", "Email", "Class", "Age", "Attendance"] \
        + [s["name"] for s in subjects] + ["Average", "Grade"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = navy

    for ri, st in enumerate(students, 2):
        marks = {m["subject_id"]: m["marks"] for m in get_student_marks(st["id"])}
        am = avg_for_student(st["id"])
        row = [st["roll_no"], st["name"], st["father"], st["email"], st["class_name"],
               st["age"], st["attendance"]] + [marks.get(s["id"], 0) for s in subjects] \
              + [am, grade_for(am)]
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"students_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/export/pdf")
@login_required(roles=["admin", "teacher"])
def export_pdf_all():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    db = get_db()
    subjects = get_subjects()
    students = db.execute("SELECT * FROM students ORDER BY name").fetchall()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Student Management PRO — Class Report", styles["Title"]), Spacer(1, 10)]

    data = [["Roll No.", "Name", "Class", "Attendance"] + [s["name"] for s in subjects] + ["Avg", "Grade"]]
    for st in students:
        marks = {m["subject_id"]: m["marks"] for m in get_student_marks(st["id"])}
        am = avg_for_student(st["id"])
        data.append([st["roll_no"], st["name"], st["class_name"] or "-", f"{st['attendance']:.0f}%"]
                    + [f"{marks.get(s['id'], 0):.0f}" for s in subjects]
                    + [f"{am:.1f}", grade_for(am)])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1F36")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    fname = f"class_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/pdf")


@app.route("/export/pdf/<int:sid>")
@login_required(roles=["admin", "teacher", "parent", "student"])
def export_pdf_one(sid):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    db = get_db()
    student = db.execute("SELECT * FROM students WHERE id=?", (sid,)).fetchone()
    if not student:
        abort(404)

    user = current_user()
    if user["role"] == "parent" and student["parent_user_id"] != user["id"]:
        abort(403)
    if user["role"] == "student" and student["student_user_id"] != user["id"]:
        abort(403)

    marks_rows = get_student_marks(sid)
    am = avg_for_student(sid)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("STUDENT REPORT CARD", styles["Title"]),
        Spacer(1, 6),
        Paragraph(f"Name: {student['name']}  |  Roll No.: {student['roll_no']}  |  "
                  f"Class: {student['class_name'] or '-'}", styles["Normal"]),
        Paragraph(f"Attendance: {student['attendance']:.0f}%", styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [["Subject", "Marks", "Grade"]]
    for r in marks_rows:
        data.append([r["name"], f"{r['marks']:.0f}", grade_for(r["marks"])])
    data.append(["Overall Average", f"{am:.1f}", grade_for(am)])

    table = Table(data, colWidths=[200, 100, 100])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1F36")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F59E0B")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    fname = f"report_{student['roll_no']}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/pdf")


# ─────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
